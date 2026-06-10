#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Chargement des fichiers OECI (format pivot, série annuelle) vers le FORMAT
INTERNE consommé par ``report_builder`` (pandas — accès aux colonnes en attribut,
ex. ``surv_df.stade``).

Principe : les fichiers OECI contiennent DÉJÀ les lignes pré-agrégées à chaque
niveau (AP-HP / GHU / Hôpital) et à chaque granularité (Total / Organe), via le
champ ``Niveau``. Ce loader SÉLECTIONNE les lignes par ``Niveau`` et les MAPPE —
il ne ré-agrège jamais (sinon double comptage). L'année est déduite du NOM de
fichier (``indicateurs_oeci_<ANNÉE>_M<MM>``).

API publique :
    charger_oeci(dossier="data", fictif=True) -> (df_aphp, df_survie)

⚠ TODO étape 4 (refonte survie) : le schéma de ``df_survie`` change ici — il
   devient long avec ``stade`` ∈ {I-III, IV} et ``population`` ∈ {tous, nouveaux}.
   Cela CASSERA ``survival_by_stage`` / ``survival_evolution`` dans
   ``chart_utils`` / ``report_builder`` (câblés sur ``stade == "II"`` et l'ancien
   schéma). NE PAS corriger ici : c'est l'objet de l'étape 4 de la migration.
"""
import os
import re
import glob

import pandas as pd
import openpyxl

# ───────────────────────── correspondances à figer ─────────────────────────
# Nom complet GHU (tel qu'écrit dans la colonne GHU des fichiers OECI) → code interne.
GHU_NOM2CODE = {
    "APHP.Centre-Université de Paris": "GHU Centre",
    "APHP.Nord-Université de Paris": "GHU Nord",
    "APHP.Hôpitaux Universitaires Henri-Mondor": "GHU Mondor",
    "APHP.Hôpitaux Universitaires Paris-Seine-Saint-Denis": "GHU PSSD",
    "APHP.Sorbonne Université": "GHU SUN",        # SUN = Sorbonne
    "APHP.Université Paris Saclay": "GHU PSL",     # PSL = Paris Saclay
}
ENTITES_APHP = {"AP-HP", "GHU Centre", "GHU Mondor", "GHU Nord", "GHU PSSD", "GHU PSL", "GHU SUN"}

# Table de correspondance EXPLICITE Niveau → (niveau d'agrégation, granularité).
# Énumérée à partir des valeurs réellement présentes dans les onglets fictifs ;
# la taxonomie est homogène entre onglets (squelette reconstruit depuis « Délais PEC »).
NIVEAU_MAP = {
    "APHP Total":     ("AP-HP", "Total"),
    "APHP Organe":    ("AP-HP", "Organe"),
    "GH Total":       ("GHU", "Total"),
    "GH Organe":      ("GHU", "Organe"),
    "Hop Total":      ("Hôpital", "Total"),
    "Hopital Organe": ("Hôpital", "Organe"),
}

# Sentinelle interne pour les lignes de sous-total (granularité Total).
SENTINELLE = "TOTAL"

RE_ANNEE = re.compile(r"indicateurs_oeci_(\d{4})_M\d{2}")

# Schémas de sortie (ordre EXACT).
COLS_APHP = [
    "annee", "entite", "appareil", "organe",
    "nb_patients", "nb_nouveaux_patients",
    "nb_sejours_chirurgie", "nb_sejours_chimiotherapie",
    "nb_sejours_radiotherapie", "nb_sejours_palliatifs",
    "delai_global_median", "delai_chirurgie_median",
    "delai_chimio_median", "delai_radio_median",
]
COMPTES_APHP = [
    "nb_patients", "nb_nouveaux_patients", "nb_sejours_chirurgie",
    "nb_sejours_chimiotherapie", "nb_sejours_radiotherapie", "nb_sejours_palliatifs",
]
MEDIANES_APHP = [
    "delai_global_median", "delai_chirurgie_median",
    "delai_chimio_median", "delai_radio_median",
]
COLS_SURVIE = [
    "annee", "entite", "appareil", "organe", "stade", "population",
    "nb_patients_stade", "survie_1an", "survie_5ans",
]

# Onglet séjour : en-tête réel → mesure interne.
MAP_SEJOUR = {
    "Séjours avec chirurgie": "nb_sejours_chirurgie",
    "Séjours avec DP chimio": "nb_sejours_chimiotherapie",
    "Séjours avec DP radioth": "nb_sejours_radiotherapie",
    "Séjours en soins palliatifs": "nb_sejours_palliatifs",
}


# ───────────────────────────── utilitaires ─────────────────────────────────
def _fichiers_oeci(dossier, fictif):
    """Liste [(annee, chemin)] triée. En mode fictif on ne lit QUE les
    ``*_fictif.xlsx`` ; en mode réel on exclut explicitement les ``_fictif`` pour
    ne jamais confondre un gabarit avec une source de données."""
    if fictif:
        motif = "indicateurs_oeci_*_M*_fictif.xlsx"
        paths = glob.glob(os.path.join(dossier, motif))
    else:
        paths = [p for p in glob.glob(os.path.join(dossier, "indicateurs_oeci_*_M*.xlsx"))
                 if "_fictif" not in os.path.basename(p)]
    out = []
    for p in sorted(paths):
        m = RE_ANNEE.search(os.path.basename(p))
        if m:
            out.append((int(m.group(1)), p))
    return sorted(out)


def _resoudre_entite(niveau_agg, ghu_nom, hop_nom):
    """AP-HP → 'AP-HP' ; GHU → code via nom complet ; Hôpital → nom d'hôpital."""
    if niveau_agg == "AP-HP":
        return "AP-HP"
    if niveau_agg == "GHU":
        return GHU_NOM2CODE.get(ghu_nom)
    return hop_nom  # Hôpital


def _ffill(ws, ligne, ncol):
    """Propage vers la droite les libellés d'un en-tête à cellules fusionnées
    (une cellule fusionnée ne porte sa valeur que dans son coin haut-gauche)."""
    vals, last = [], None
    for c in range(1, ncol + 1):
        v = ws.cell(ligne, c).value
        if v not in (None, ""):
            last = v
        vals.append(last)
    return vals


def _normaliser_dims(df, garder):
    """Ajoute entite/appareil/organe normalisés à un DataFrame dont les 5 premières
    colonnes sont (Niveau, GHU, Hôpital, Appareil, Organe). ``garder`` = ensemble
    des niveaux d'agrégation conservés. Applique la sentinelle TOTAL en granularité
    Total (piloté par le Niveau, pas par le contenu des cellules)."""
    cols = list(df.columns)
    df = df.rename(columns={cols[0]: "niveau", cols[1]: "ghu",
                            cols[2]: "hopital", cols[3]: "appareil", cols[4]: "organe"})
    df = df[df["niveau"].isin(NIVEAU_MAP)].copy()
    df["niveau_agg"] = df["niveau"].map(lambda n: NIVEAU_MAP[n][0])
    df["granularite"] = df["niveau"].map(lambda n: NIVEAU_MAP[n][1])
    df = df[df["niveau_agg"].isin(garder)].copy()
    df["entite"] = [_resoudre_entite(a, g, h)
                    for a, g, h in zip(df["niveau_agg"], df["ghu"], df["hopital"])]
    df = df[df["entite"].notna()].copy()
    est_total = df["granularite"].eq("Total")
    df.loc[est_total, "appareil"] = SENTINELLE
    df.loc[est_total, "organe"] = SENTINELLE
    return df


# ──────────────────────────── lecture par onglet ───────────────────────────
def _frame_patient(path, annee):
    df = pd.read_excel(path, sheet_name="Indicateurs patient")
    df.columns = [c.strip() if isinstance(c, str) else c for c in df.columns]
    df = _normaliser_dims(df, garder={"AP-HP", "GHU"})
    df = df.rename(columns={"Nb patients": "nb_patients", "Nvx patients": "nb_nouveaux_patients"})
    df["annee"] = annee
    return df[["annee", "entite", "appareil", "organe", "nb_patients", "nb_nouveaux_patients"]]


def _frame_sejour(path, annee):
    df = pd.read_excel(path, sheet_name="Indicateurs séjour")
    df.columns = [c.strip() if isinstance(c, str) else c for c in df.columns]
    df = _normaliser_dims(df, garder={"AP-HP", "GHU"})
    df = df.rename(columns=MAP_SEJOUR)
    df["annee"] = annee
    return df[["annee", "entite", "appareil", "organe"] + list(MAP_SEJOUR.values())]


def _bloc_delai_vers_champ(label):
    """Libellé de bloc « Délais PEC » → mesure interne (tolérant aux variantes
    d'orthographe du gabarit : MEDECINE, RAFIOTHERAPIE…)."""
    u = (label or "").upper()
    if "TOTAL" in u:
        return "delai_global_median"
    if "CHIR" in u:
        return "delai_chirurgie_median"
    if "MED" in u or "MÉDEC" in u:        # MÉDECINE ↔ chimiothérapie
        return "delai_chimio_median"
    if "RADIO" in u or "RAFIO" in u or "THERAP" in u:
        return "delai_radio_median"
    return None


def _colonnes_mediane_delais(ws):
    """{mesure interne → index de colonne (1-based)} pour les « Médiane délais »
    de chaque bloc (TOTAL/CHIRURGIE/MÉDECINE/RADIOTHÉRAPIE), via l'en-tête 2 lignes."""
    ncol = ws.max_column
    blocs = _ffill(ws, 1, ncol)                                   # libellé de bloc propagé
    sous = [ws.cell(2, c).value for c in range(1, ncol + 1)]      # sous-colonne (ligne 2)
    res = {}
    for c in range(1, ncol + 1):
        if sous[c - 1] and "édiane" in str(sous[c - 1]):         # « Médiane délais »
            champ = _bloc_delai_vers_champ(blocs[c - 1])
            if champ:
                res[champ] = c
    return res


def _frame_delais(path, annee):
    # NB : chargement normal (pas read_only) — l'accès aléatoire ws.cell(r, c) est
    # O(n) par appel en mode read_only et rend la lecture pathologiquement lente.
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Délais PEC"]
    colmap = _colonnes_mediane_delais(ws)   # dims : 1 Niveau,2 GHU,3 Hôpital,4 Appareil,5 Organe
    lignes = []
    for r in range(3, ws.max_row + 1):       # données à partir de la ligne 3
        niv = ws.cell(r, 1).value
        if niv not in NIVEAU_MAP:
            continue
        agg, gran = NIVEAU_MAP[niv]
        if agg not in ("AP-HP", "GHU"):
            continue
        ent = _resoudre_entite(agg, ws.cell(r, 2).value, ws.cell(r, 3).value)
        if ent is None:
            continue
        app = SENTINELLE if gran == "Total" else ws.cell(r, 4).value
        org = SENTINELLE if gran == "Total" else ws.cell(r, 5).value
        ligne = {"annee": annee, "entite": ent, "appareil": app, "organe": org}
        ligne.update({champ: ws.cell(r, c).value for champ, c in colmap.items()})
        lignes.append(ligne)
    wb.close()
    return pd.DataFrame(lignes)


# ─────────────────────────────── survie ────────────────────────────────────
def _norm_population(label):
    u = str(label or "")
    if "Tous" in u:
        return "tous"
    if "Nouveau" in u:
        return "nouveaux"
    return None


def _norm_horizon(label):
    u = str(label or "")
    if "5 ans" in u:
        return "5ans"
    if "1 an" in u:
        return "1an"
    return None


def _norm_stade(label):
    u = str(label or "")
    if "I-III" in u:
        return "I-III"
    if "IV" in u:
        return "IV"
    return None


def _plan_survie(ws):
    """Plan des colonnes mesures de « Survie globale » (en-tête 4 lignes) :
    {(population, stade) → {'nb': col, 's5': col, 's1': col}}.
    Le « Nb patients » est identique pour 5 ans et 1 an (même effectif de stade) :
    on retient celui du bloc 5 ans."""
    ncol = ws.max_column
    pop = _ffill(ws, 1, ncol)
    hor = _ffill(ws, 2, ncol)
    sta = _ffill(ws, 3, ncol)
    kind = [ws.cell(4, c).value for c in range(1, ncol + 1)]
    plan = {}
    for c in range(6, ncol + 1):
        p, h, s = _norm_population(pop[c - 1]), _norm_horizon(hor[c - 1]), _norm_stade(sta[c - 1])
        if p is None or h is None or s is None:
            continue
        d = plan.setdefault((p, s), {})
        if "survie" in str(kind[c - 1] or "").lower():   # « % survie »
            d["s5" if h == "5ans" else "s1"] = c
        elif h == "5ans":                                # « Nb patients » (bloc 5 ans)
            d["nb"] = c
    return plan


def _frame_survie(path, annee):
    """Déplie les blocs survie vers le format long. Dimensions de l'onglet :
    1 Niveau, 2 Appareil, 3 Organe, 4 GHU, 5 Hôpital. Niveaux conservés : tous
    (AP-HP + GHU + Hôpital), données à partir de la ligne 5."""
    wb = openpyxl.load_workbook(path, data_only=True)   # accès aléatoire : pas de read_only
    ws = wb["Survie globale"]
    plan = _plan_survie(ws)
    lignes = []
    for r in range(5, ws.max_row + 1):
        niv = ws.cell(r, 1).value
        if niv not in NIVEAU_MAP:
            continue
        agg, gran = NIVEAU_MAP[niv]
        ent = _resoudre_entite(agg, ws.cell(r, 4).value, ws.cell(r, 5).value)
        if ent is None:
            continue
        app = SENTINELLE if gran == "Total" else ws.cell(r, 2).value
        org = SENTINELLE if gran == "Total" else ws.cell(r, 3).value
        for (pop, stade), cols in plan.items():
            lignes.append({
                "annee": annee, "entite": ent, "appareil": app, "organe": org,
                "stade": stade, "population": pop,
                "nb_patients_stade": ws.cell(r, cols["nb"]).value,
                "survie_1an": ws.cell(r, cols["s1"]).value,
                "survie_5ans": ws.cell(r, cols["s5"]).value,
            })
    wb.close()
    return pd.DataFrame(lignes)


# ──────────────────────────── assemblage public ────────────────────────────
def _assembler_aphp(fichiers):
    pat = pd.concat([_frame_patient(p, a) for a, p in fichiers], ignore_index=True)
    sej = pd.concat([_frame_sejour(p, a) for a, p in fichiers], ignore_index=True)
    dele = pd.concat([_frame_delais(p, a) for a, p in fichiers], ignore_index=True)
    cle = ["annee", "entite", "appareil", "organe"]
    df = pat.merge(sej, on=cle, how="outer").merge(dele, on=cle, how="outer")
    # Comptages absents = 0 réel ; médiane absente (organe sans données délais) = 0
    # sentinelle (jamais NaN, contrat report_builder).
    df[COMPTES_APHP] = df[COMPTES_APHP].fillna(0)
    df[MEDIANES_APHP] = df[MEDIANES_APHP].fillna(0.0)
    for c in COMPTES_APHP:
        df[c] = df[c].round().astype(int)
    for c in MEDIANES_APHP:
        df[c] = df[c].astype(float)
    return df[COLS_APHP].sort_values(cle).reset_index(drop=True)


def charger_oeci(dossier="data", fictif=True):
    """Charge la série annuelle OECI → (df_aphp, df_survie) au format interne.

    df_aphp : une ligne par (annee, entite, appareil, organe) ; entite ∈ AP-HP + 6 GHU.
    df_survie : format long (annee, entite, appareil, organe, stade, population) ;
                entite ∈ AP-HP + GHU + Hôpital.
    """
    fichiers = _fichiers_oeci(dossier, fictif)
    if not fichiers:
        mode = "fictif" if fictif else "réel"
        raise FileNotFoundError(
            f"Aucun fichier OECI ({mode}) trouvé dans {dossier!r} "
            f"(motif indicateurs_oeci_AAAA_MMM{'_fictif' if fictif else ''}.xlsx).")
    df_aphp = _assembler_aphp(fichiers)
    df_survie = pd.concat([_frame_survie(p, a) for a, p in fichiers], ignore_index=True)
    df_survie = df_survie[COLS_SURVIE].reset_index(drop=True)
    return df_aphp, df_survie


# ─────────────────────────── validation standalone ─────────────────────────
def _valider(df_aphp, df_survie):
    pb = []

    # 1) df_aphp : colonnes EXACTES, entite attendue, aucune NaN dans les mesures.
    if list(df_aphp.columns) != COLS_APHP:
        pb.append(f"df_aphp colonnes ≠ schéma : {list(df_aphp.columns)}")
    ent_inc = set(df_aphp["entite"]) - ENTITES_APHP
    if ent_inc:
        pb.append(f"df_aphp entite inattendues : {ent_inc}")
    mesures = COMPTES_APHP + MEDIANES_APHP
    if int(df_aphp[mesures].isna().sum().sum()) != 0:
        pb.append("df_aphp : NaN présentes dans les mesures")

    # 2) Identité d'agrégation issue DES DONNÉES : AP-HP == Σ GHU, par (annee, appareil, organe).
    ecarts = {}
    for mesure in ("nb_patients", "nb_sejours_chirurgie"):
        piv = df_aphp.pivot_table(index=["annee", "appareil", "organe"],
                                  columns="entite", values=mesure,
                                  aggfunc="sum", fill_value=0)
        ghu_cols = [c for c in piv.columns if c != "AP-HP"]
        aphp = piv["AP-HP"] if "AP-HP" in piv.columns else 0
        ecart = (aphp - piv[ghu_cols].sum(axis=1)).abs().max()
        ecarts[mesure] = int(ecart)
        if ecart != 0:
            pb.append(f"identité AP-HP==ΣGHU violée pour {mesure} (écart max {ecart})")

    # 3) df_survie : schéma, domaines, ordres de survie au niveau AP-HP.
    if list(df_survie.columns) != COLS_SURVIE:
        pb.append(f"df_survie colonnes ≠ schéma : {list(df_survie.columns)}")
    if set(df_survie["stade"]) - {"I-III", "IV"}:
        pb.append(f"df_survie stade hors domaine : {set(df_survie['stade'])}")
    if set(df_survie["population"]) - {"tous", "nouveaux"}:
        pb.append(f"df_survie population hors domaine : {set(df_survie['population'])}")
    # I-III > IV (5 ans) et 1 an > 5 ans (I-III), sur AP-HP
    ap = df_survie[df_survie["entite"] == "AP-HP"]
    larg = ap.pivot_table(index=["annee", "appareil", "organe", "population"],
                          columns="stade", values="survie_5ans")
    if {"I-III", "IV"}.issubset(larg.columns):
        viol = int((larg["I-III"] <= larg["IV"]).sum())
        if viol:
            pb.append(f"survie AP-HP : {viol} cas I-III ≤ IV (5 ans)")
    i3 = ap[ap["stade"] == "I-III"]
    viol_h = int((i3["survie_1an"] <= i3["survie_5ans"]).sum())
    if viol_h:
        pb.append(f"survie AP-HP : {viol_h} cas 1 an ≤ 5 ans (I-III)")

    return pb, ecarts


def _main():
    print("Chargement OECI fictif (data/)…")
    df_aphp, df_survie = charger_oeci("data", fictif=True)
    pb, ecarts = _valider(df_aphp, df_survie)

    print("\n=== df_aphp ===")
    print(f"  lignes : {len(df_aphp):,}")
    print(f"  années : {sorted(df_aphp['annee'].unique())}")
    print(f"  entites : {sorted(df_aphp['entite'].unique())}")
    print(f"  organes distincts : {df_aphp['organe'].nunique()} (dont sentinelle {SENTINELLE!r})")
    print("  exemple AP-HP Total par année :")
    ex = df_aphp[(df_aphp.entite == "AP-HP") & (df_aphp.organe == SENTINELLE)]
    for _, row in ex.iterrows():
        print(f"    {row.annee} : {row.nb_patients:>7,} patients | "
              f"chir {row.nb_sejours_chirurgie:>6,} | délai global méd. {row.delai_global_median}")

    print("\n=== df_survie ===")
    print(f"  lignes : {len(df_survie):,}")
    print(f"  années : {sorted(df_survie['annee'].unique())}")
    print(f"  entites (niveaux présents) : {sorted(df_survie['entite'].unique())}")
    print(f"  stade : {sorted(df_survie['stade'].unique())} | "
          f"population : {sorted(df_survie['population'].unique())}")
    aphp_tot = df_survie[(df_survie.entite == "AP-HP") & (df_survie.organe == SENTINELLE)
                         & (df_survie.population == "tous")]
    print("  survie AP-HP (toutes localisations, pop=tous) par année :")
    for annee in sorted(aphp_tot["annee"].unique()):
        sub = aphp_tot[aphp_tot.annee == annee]
        s13 = sub[sub.stade == "I-III"].iloc[0]
        s4 = sub[sub.stade == "IV"].iloc[0]
        print(f"    {annee} : I-III 5ans={s13.survie_5ans} 1an={s13.survie_1an} | "
              f"IV 5ans={s4.survie_5ans} 1an={s4.survie_1an}")

    print("\n=== Validation ===")
    print(f"  identité AP-HP==ΣGHU : écarts max {ecarts}")
    if pb:
        print(f"  {len(pb)} ANOMALIE(S) :")
        for m in pb:
            print(f"    - {m}")
    else:
        print("  OK — schémas, domaines, agrégation et ordres de survie conformes.")


if __name__ == "__main__":
    _main()
