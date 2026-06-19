#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Chargeur RÉEL piloté par le YAML — ingestion des vrais fichiers sources vers le
format pivot LONG unique (``data/donnees.csv``).

Seul point d'ingestion des fichiers sources réels : un moteur générique unique,
dont la STRUCTURE des fichiers (feuilles, positions de dimensions,
en-têtes, taxonomie de niveau, forme GHU, coercition) est déclarée dans
``docs/descriptif_sources.yaml`` (cf. ``contrat_donnees_pivot.md``). Le **mapping
colonne/bloc/axe brut → variable interne** reste EN CODE (tables ci-dessous), keyé
sur les libellés du YAML.

Dispositions de mesures gérées : ``simple`` (une mesure par colonne), ``blocs``
(délais : bloc ligne 1 × sous-colonne ligne 2), ``plan_survie`` (population × horizon
× stade, en-tête 4 lignes).

Sortie : DataFrame long au schéma du contrat (``format_long.LONG_COLS``). Les
transforms procéduraux restent EN CODE : filtre de période, reconstruction de la
survie niveau appareil (pondération NaN-safe), sentinelles ``TOTAL``.

API publique :
    charger_long(dossier="data", fictif=False) -> pd.DataFrame   (format long)
    mapping_hopital_ghu(...) / mapping_hopital_ghu_delais(...)    (hôpital → code GHU)
"""
import os
import re
import glob

import numpy as np
import pandas as pd
import openpyxl
import yaml

from referentiels import (GHU_NOM2CODE, STATUT2TYPE, ANNEE_MIN, ANNEE_MAX)
from format_long import LONG_COLS, niveau_depuis_entite

_DESCRIPTIF = os.path.join(os.path.dirname(__file__), "..", "docs", "descriptif_sources.yaml")
SENTINELLE = "TOTAL"

# ── Mappings EN CODE (libellé brut du YAML → variable interne / population) ──────
# OECI « Indicateurs patient » : colonne → (variable, population).
MAP_OECI_PATIENT = {
    "Nb patients":  ("nb_patients", "tous"),
    "Nvx patients": ("nb_patients", "nouveaux"),
}
# OECI « Indicateurs séjour » : colonne → variable (population=tous).
MAP_OECI_SEJOUR = {
    "Séjours avec chirurgie":      "nb_sejours_chirurgie",
    "Séjours avec DP chimio":      "nb_sejours_chimiotherapie",
    "Séjours avec DP radioth":     "nb_sejours_radiotherapie",
    "Séjours en soins palliatifs": "nb_sejours_palliatifs",
}
# Régional « Total » patients : colonne → (variable, population).
MAP_BN_PATIENT = {
    "Nb de patients":    ("nb_patients", "tous"),
    "Nouveaux patients": ("nb_patients", "nouveaux"),
}
# Régional « Total » séjours : colonne → variable (population=tous). « Séjours avec
# chirurgie » apparaît 2× (0 jour / > 0 jour) → SOMMÉES.
MAP_BN_SEJOUR = {
    "Séjours avec chirurgie":     "nb_sejours_chirurgie",
    "Séjours avec chimio":        "nb_sejours_chimiotherapie",
    "Séjours avec radiothérapie": "nb_sejours_radiotherapie",
    "Nb de séjours palliatifs":   "nb_sejours_palliatifs",
}


def _bloc_delai_vers_variable(label):
    """Bloc « Délais PEC » (ligne 1) → variable interne, par MOTS-CLÉS (tolérant aux
    fautes de frappe source : MEDECINE = traitement médical, RAFIOTHERAPIE = radio)."""
    u = (label or "").upper()
    if "TOTAL" in u:
        return "delai_global_median"
    if "CHIR" in u:
        return "delai_chirurgie_median"
    if "MED" in u or "MÉDEC" in u:                       # MEDECINE = oncologie médicale
        return "delai_traitement_medical_median"
    if "RADIO" in u or "RAFIO" in u or "THERAP" in u:
        return "delai_radio_median"
    return None


# ── Référentiel YAML ────────────────────────────────────────────────────────────
def _charger_descriptif():
    with open(_DESCRIPTIF, encoding="utf-8") as f:
        return yaml.safe_load(f)


# ── Coercition numérique (format FR ; jetons masqués → NaN) ──────────────────────
def _coercer_valeur(brut, masque):
    """Une cellule brute → float (NaN si vide/masquée). Retire espaces normaux et
    insécables + « % », virgule décimale → point. Robuste au format français réel."""
    if brut is None:
        return np.nan
    if isinstance(brut, (int, float)) and not isinstance(brut, bool):
        return float(brut)
    s = str(brut).strip()
    if s == "" or s in masque or re.fullmatch(r"\.+", s):
        return np.nan
    s = s.replace(" ", "").replace(" ", "").replace("%", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return np.nan


# ── Résolution niveau / entité ──────────────────────────────────────────────────
def _code_ghu(nom):
    """Code GHU depuis une forme longue ou courte ; tolère le séparateur '.'/',' des
    formes courtes selon les millésimes."""
    if nom is None:
        return None
    cle = str(nom).strip()
    return GHU_NOM2CODE.get(cle) or GHU_NOM2CODE.get(cle.replace(",", "."))


def _resoudre_niveau(niveau_val, conf_niveau):
    """(agg, gran) d'une valeur de Niveau selon la conf YAML. ``gran`` vaut None en
    mode ``cellules`` (granularité déduite du remplissage Appareil/Organe)."""
    if isinstance(conf_niveau, dict) and conf_niveau.get("mode") == "mots_cles":
        u = str(niveau_val or "").upper()
        # 'HOP' avant 'GH'/'AP-HP' (cf. libellés EDS).
        for cle, agg in (("HOP", "hopital"), ("HÔP", "hopital"),
                         ("GH", "ghu"), ("AP-HP", "aphp"), ("APHP", "aphp")):
            if cle in u:
                return agg, None
        return None, None
    # Sinon : map de libellés exacts → {agg, gran}.
    entree = conf_niveau.get(niveau_val) if isinstance(conf_niveau, dict) else None
    if entree is None:
        return None, None
    return entree.get("agg"), entree.get("gran")


def _resoudre_entite_oeci(agg, ghu_val, hop_val):
    """OECI : aphp → 'AP-HP' ; ghu → code via nom (long/court) ; hopital → nom brut."""
    if agg == "aphp":
        return "AP-HP"
    if agg == "ghu":
        return _code_ghu(ghu_val)
    if agg == "hopital":
        return None if hop_val in (None, "") else str(hop_val).strip()
    return None


# ── En-têtes multi-lignes ────────────────────────────────────────────────────────
def _ffill_ligne(ws, ligne, ncol):
    """Propage à droite les libellés d'un en-tête à cellules fusionnées."""
    vals, last = [], None
    for c in range(1, ncol + 1):
        v = ws.cell(ligne, c).value
        if v not in (None, ""):
            last = v
        vals.append(last)
    return vals


# ── Moteur OECI (un fichier par année) ──────────────────────────────────────────
def _colonnes_simple(ws, ligne_entete):
    """{libellé de colonne (strip) → [index 1-based]} (liste : doublons possibles)."""
    cols = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(ligne_entete, c).value
        if v not in (None, ""):
            cols.setdefault(str(v).strip(), []).append(c)
    return cols


def _mesures_simple_oeci(ws, r, cols_idx, masque):
    """Disposition ``simple`` OECI : (variable, population, valeur) par colonne utile.
    ``cols_idx`` = {variable_or_patientcol → indices}. On gère patient (tous/nouveaux)
    et séjour via les deux tables de mapping."""
    out = []
    for brut, idxs in cols_idx.items():
        if brut in MAP_OECI_PATIENT:
            var, pop = MAP_OECI_PATIENT[brut]
        elif brut in MAP_OECI_SEJOUR:
            var, pop = MAP_OECI_SEJOUR[brut], "tous"
        else:
            continue
        val = _coercer_valeur(ws.cell(r, idxs[0]).value, masque)
        out.append((var, pop, None, val))
    return out


def _colonnes_delais(ws, masque):
    """Disposition ``blocs`` : {variable délai → index de la « Médiane délais » du
    bloc} via l'en-tête 2 lignes (L1 = bloc propagé, L2 = sous-colonne)."""
    ncol = ws.max_column
    blocs = _ffill_ligne(ws, 1, ncol)
    sous = [ws.cell(2, c).value for c in range(1, ncol + 1)]
    res = {}
    for c in range(1, ncol + 1):
        if sous[c - 1] and "édiane" in str(sous[c - 1]):
            var = _bloc_delai_vers_variable(blocs[c - 1])
            if var:
                res[var] = c
    return res


def _plan_survie(ws):
    """Disposition ``plan_survie`` : {(population, stade) → {'nb','s5','s1' cols}}.
    En-tête 4 lignes : L1 population, L2 horizon, L3 stade, L4 type (Nb/% survie)."""
    ncol = ws.max_column
    pop = _ffill_ligne(ws, 1, ncol)
    hor = _ffill_ligne(ws, 2, ncol)
    sta = _ffill_ligne(ws, 3, ncol)
    kind = [ws.cell(4, c).value for c in range(1, ncol + 1)]

    def _np(l):   # population
        u = str(l or "")
        return "tous" if "Tous" in u else ("nouveaux" if "Nouveau" in u else None)

    def _nh(l):   # horizon
        u = str(l or "")
        return "5ans" if "5 ans" in u else ("1an" if "1 an" in u else None)

    def _ns(l):   # stade
        u = str(l or "")
        return "I-III" if "I-III" in u else ("IV" if "IV" in u else None)

    plan = {}
    for c in range(6, ncol + 1):
        p, h, s = _np(pop[c - 1]), _nh(hor[c - 1]), _ns(sta[c - 1])
        if p is None or h is None or s is None:
            continue
        d = plan.setdefault((p, s), {})
        if "survie" in str(kind[c - 1] or "").lower():
            d["s5" if h == "5ans" else "s1"] = c
        elif h == "5ans":
            d["nb"] = c
    return plan


def _lire_feuille_oeci(ws, conf, source, annee, masque):
    """Lit une feuille OECI selon sa conf YAML → liste de lignes longues (dicts)."""
    dims = conf["dimensions"]
    conf_niv = conf["niveau"]
    disp = conf["mesures"]["disposition"]
    r0 = conf["premiere_ligne_donnees"]
    granularite = conf.get("granularite")

    # Index de colonnes selon la disposition.
    if disp == "simple":
        cols_idx = _colonnes_simple(ws, conf["lignes_entete"])
    elif disp == "blocs":
        delais_cols = _colonnes_delais(ws, masque)
    elif disp == "plan_survie":
        plan = _plan_survie(ws)

    lignes = []
    for r in range(r0, ws.max_row + 1):
        niv_val = ws.cell(r, dims["niveau"]).value
        agg, gran = _resoudre_niveau(niv_val, conf_niv)
        if agg is None:
            continue
        ghu_val = ws.cell(r, dims["ghu"]).value if "ghu" in dims else None
        hop_val = ws.cell(r, dims["hopital"]).value if "hopital" in dims else None
        ent = _resoudre_entite_oeci(agg, ghu_val, hop_val)
        if ent is None:
            continue
        app_cell = ws.cell(r, dims["appareil"]).value
        org_cell = ws.cell(r, dims["organe"]).value
        # Granularité : par le Niveau (gran) ou par les cellules (mode 'cellules').
        if granularite == "cellules":
            total = app_cell in (None, "") and org_cell in (None, "")
        else:
            total = (gran == "total")
        app = SENTINELLE if total else app_cell
        org = SENTINELLE if total else org_cell

        base = dict(annee=annee, source=source, niveau=agg, entite=ent,
                    appareil=app, organe=org, age="tous")
        if disp == "simple":
            for var, pop, stade, val in _mesures_simple_oeci(ws, r, cols_idx, masque):
                lignes.append({**base, "stade": stade, "population": pop,
                               "variable": var, "valeur": val})
        elif disp == "blocs":
            for var, c in delais_cols.items():
                lignes.append({**base, "stade": None, "population": "tous",
                               "variable": var,
                               "valeur": _coercer_valeur(ws.cell(r, c).value, masque)})
        elif disp == "plan_survie":
            for (pop, stade), cc in plan.items():
                for var, key in (("nb_patients_stade", "nb"),
                                 ("survie_1an", "s1"), ("survie_5ans", "s5")):
                    if key in cc:
                        lignes.append({**base, "stade": stade, "population": pop,
                                       "variable": var,
                                       "valeur": _coercer_valeur(ws.cell(r, cc[key]).value, masque)})
    return lignes


def _fichiers_oeci(dossier, motif, fictif):
    paths = glob.glob(os.path.join(dossier, motif))
    if fictif:
        paths = [p for p in paths if p.endswith("_fictif.xlsx")]
    else:
        paths = [p for p in paths if "_fictif" not in os.path.basename(p)]
    out = []
    for p in sorted(paths):
        m = re.search(r"indicateurs_oeci_(\d{4})_M\d{2}", os.path.basename(p))
        if m:
            out.append((int(m.group(1)), p))
    return sorted(out)


def _charger_oeci_long(dossier, conf_oeci, fictif):
    """Toutes les feuilles OECI consommées, tous millésimes → lignes longues."""
    fichiers = _fichiers_oeci(dossier, conf_oeci["fichier"], fictif)
    if not fichiers:
        mode = "fictif" if fictif else "réel"
        raise FileNotFoundError(f"Aucun fichier OECI ({mode}) dans {dossier!r}.")
    masque = set(conf_oeci.get("coercition", {}).get("masque", []))
    feuilles = [(nom, c) for nom, c in conf_oeci["feuilles"].items()
                if c.get("consommee", True) and "mesures" in c]
    lignes = []
    for annee, path in fichiers:
        wb = openpyxl.load_workbook(path, data_only=True)  # cf. CLAUDE.md : pas de read_only
        for nom, c in feuilles:
            if c["nom"] not in wb.sheetnames:
                if c.get("obligatoire"):
                    print(f"  ⚠ OECI {annee} : feuille obligatoire « {c['nom']} » absente.")
                continue
            lignes += _lire_feuille_oeci(wb[c["nom"]], c, c["source"], annee, masque)
        wb.close()
    return lignes


# ── Moteur régional (BN — multi-années, colonne Année) ──────────────────────────
def _entite_regional(hopital_aphp, statut):
    if str(hopital_aphp).strip() == "Oui":
        return "AP-HP"
    return STATUT2TYPE.get(str(statut).strip())


def _lire_feuille_regional(path, conf, source, masque):
    """Feuille régionale « Total » → lignes longues. Multi-Finess agrégées par TYPE
    d'établissement (somme). « Séjours avec chirurgie » apparaît 2× → sommées."""
    dims = conf["dimensions"]
    header = conf["lignes_entete"] - 1                       # pandas: 0-based
    df = pd.read_excel(path, sheet_name=conf["nom"], header=header)
    df.columns = [str(c).strip() for c in df.columns]
    # Dimensions par POSITION.
    pos = {k: v - 1 for k, v in dims.items()}
    niveau = df.iloc[:, pos["niveau"]]
    conf_niv = conf["niveau"]
    keep = niveau.isin(list(conf_niv))
    df = df[keep].copy()
    niveau = df.iloc[:, pos["niveau"]]
    appareil = df.iloc[:, pos["appareil"]].astype(object)
    organe = df.iloc[:, pos["organe"]].astype(object)
    annee = pd.to_numeric(df.iloc[:, pos["annee"]], errors="coerce")
    entite = [_entite_regional(h, s) for h, s
              in zip(df.iloc[:, pos["hopital_aphp"]], df.iloc[:, pos["statut"]])]
    # Sentinelle pilotée par le Niveau.
    gran = niveau.map(lambda n: conf_niv[n]["gran"])
    app = appareil.where(gran != "total", SENTINELLE)
    org = organe.where(gran.isin(["total", "appareil"]), organe)
    org = org.where(~gran.isin(["total", "appareil"]), SENTINELLE)

    cadre = pd.DataFrame({"annee": annee, "entite": entite,
                          "appareil": app.values, "organe": org.values})
    def _indices(brut):
        # Colonnes valant ``brut`` OU ``brut.N`` (suffixe de déduplication pandas pour
        # les en-têtes répétés, ex. « Séjours avec chirurgie » sous « 0 jour » ET
        # « > 0 jour ») → toutes SOMMÉES.
        return [i for i, c in enumerate(df.columns)
                if c == brut or re.fullmatch(re.escape(brut) + r"\.\d+", str(c))]

    # Mesures (somme par ligne pour les colonnes en double, ex. chirurgie ×2).
    for brut, var in MAP_BN_SEJOUR.items():
        idxs = _indices(brut)
        if idxs:
            cadre[var] = sum(df.iloc[:, i].map(lambda v: _coercer_valeur(v, masque))
                             for i in idxs)
    for brut, (var, pop) in MAP_BN_PATIENT.items():
        idxs = _indices(brut)
        if idxs:
            cadre[(var, pop)] = df.iloc[:, idxs[0]].map(lambda v: _coercer_valeur(v, masque))

    cadre = cadre[cadre["entite"].notna() & cadre["annee"].notna()].copy()
    cadre["annee"] = cadre["annee"].round().astype(int)

    # Agrégation par TYPE (somme) puis émission longue.
    mesure_cols = [c for c in cadre.columns if c not in ("annee", "entite", "appareil", "organe")]
    agg = cadre.groupby(["annee", "entite", "appareil", "organe"], as_index=False)[mesure_cols].sum()
    lignes = []
    for _, row in agg.iterrows():
        base = dict(annee=int(row["annee"]), source=source,
                    niveau=niveau_depuis_entite(row["entite"]), entite=row["entite"],
                    appareil=row["appareil"], organe=row["organe"], age="tous", stade=None)
        for col in mesure_cols:
            var, pop = col if isinstance(col, tuple) else (col, "tous")
            lignes.append({**base, "population": pop, "variable": var, "valeur": row[col]})
    return lignes


def _trouver_regional(dossier, motif, fictif):
    cands = glob.glob(os.path.join(dossier, motif))
    cands = [p for p in cands if (p.endswith("_fictif.xlsx") == fictif)]
    if not cands:
        mode = "fictif" if fictif else "réel"
        raise FileNotFoundError(f"Aucun fichier régional ({mode}) : {motif}")
    return sorted(cands)[0]


def _charger_regional_long(dossier, conf_reg, fictif):
    masque = set(conf_reg.get("coercition", {}).get("masque", []))
    source = conf_reg["source"]
    lignes = []
    for nom, c in conf_reg["feuilles"].items():
        motif = conf_reg["fichiers"][c["fichier"]]
        path = _trouver_regional(dossier, motif, fictif)
        lignes += _lire_feuille_regional(path, c, source, masque)
    return lignes


# ── Transforms procéduraux ──────────────────────────────────────────────────────
def _reconstruire_survie_appareil(long):
    """Reconstruit la survie niveau APPAREIL (organe=TOTAL) par agrégation pondérée
    NaN-safe des organes (clé incluant population). Émet des lignes longues EDS APHP.
    Calque de ``export_internes._survie_niveau_appareil`` sur le format long."""
    surv = long[(long["source"] == "EDS APHP") & (long["organe"] != SENTINELLE)]
    if surv.empty:
        return long
    wide = surv.pivot_table(
        index=["annee", "entite", "appareil", "organe", "stade", "population"],
        columns="variable", values="valeur", aggfunc="first").reset_index()
    cle = ["annee", "entite", "appareil", "stade", "population"]
    nb = pd.to_numeric(wide.get("nb_patients_stade"), errors="coerce")

    def _round1(serie):
        return serie.map(lambda x: round(float(x), 1) if pd.notna(x) else x)

    for mesure, wm, wp in (("survie_1an", "_wm1", "_p1"), ("survie_5ans", "_wm5", "_p5")):
        m = pd.to_numeric(wide.get(mesure), errors="coerce")
        valide = m.notna() & nb.notna()
        wide[wp] = nb.fillna(0) * valide
        wide[wm] = m.fillna(0) * wide[wp]
    g = wide.groupby(cle, as_index=False).agg(
        nb_patients_stade=("nb_patients_stade", "sum"),
        _wm1=("_wm1", "sum"), _p1=("_p1", "sum"),
        _wm5=("_wm5", "sum"), _p5=("_p5", "sum"))
    g["survie_1an"] = _round1(g["_wm1"] / g["_p1"].where(g["_p1"] > 0))
    g["survie_5ans"] = _round1(g["_wm5"] / g["_p5"].where(g["_p5"] > 0))
    g["organe"] = SENTINELLE
    lignes = []
    for _, row in g.iterrows():
        base = dict(annee=int(row["annee"]), source="EDS APHP",
                    niveau=niveau_depuis_entite(row["entite"]), entite=row["entite"],
                    appareil=row["appareil"], organe=SENTINELLE, age="tous",
                    stade=row["stade"], population=row["population"])
        for var in ("nb_patients_stade", "survie_1an", "survie_5ans"):
            if pd.notna(row[var]):           # pas de bourrage NaN (survie tout masquée)
                lignes.append({**base, "variable": var, "valeur": row[var]})
    return pd.concat([long, pd.DataFrame(lignes)], ignore_index=True)


def _filtrer_periode(long):
    annee = pd.to_numeric(long["annee"], errors="coerce")
    return long[(annee >= ANNEE_MIN) & (annee <= ANNEE_MAX)].copy()


# Variables attendues (var, population) à COMPLÉTER à 0 par clé (transform procédural
# de l'ancien pipeline : merge externe patient/séjour/délais + fillna(0)). Le grain
# hôpital (comptes nouveaux) et la survie n'en font PAS partie.
_ATTENDUS_DIM = [
    ("nb_patients", "tous"), ("nb_patients", "nouveaux"),
    ("nb_sejours_chirurgie", "tous"), ("nb_sejours_chimiotherapie", "tous"),
    ("nb_sejours_radiotherapie", "tous"), ("nb_sejours_palliatifs", "tous"),
    ("delai_global_median", "tous"), ("delai_chirurgie_median", "tous"),
    ("delai_traitement_medical_median", "tous"), ("delai_radio_median", "tous"),
]
_ATTENDUS_BN = _ATTENDUS_DIM[:6]   # régional : pas de délais


def _completer_zeros(long, source, niveaux, attendus):
    """Complète à 0 les ``attendus`` (var, population) sur l'UNION des clés
    (annee, entite, appareil, organe) d'un (source, niveaux) — reproduit le merge
    externe + fillna(0) de l'ancien pipeline (les comptes/délais consommés ne sont
    jamais NaN côté report_builder)."""
    masque = (long["source"] == source) & (long["niveau"].isin(niveaux))
    sub, autres = long[masque], long[~masque]
    if sub.empty:
        return long
    sub = sub.copy()
    sub["vp"] = list(zip(sub["variable"], sub["population"]))
    sub["org"] = sub["organe"].fillna("\x00")
    wide = sub.pivot_table(index=["annee", "entite", "appareil", "org"],
                           columns="vp", values="valeur", aggfunc="first")
    wide = wide.reindex(columns=attendus).fillna(0.0).reset_index()
    lignes = []
    for _, row in wide.iterrows():
        organe = row["org"]
        organe = np.nan if organe == "\x00" else organe
        base = dict(annee=int(row["annee"]), source=source,
                    niveau=niveau_depuis_entite(row["entite"]), entite=row["entite"],
                    appareil=row["appareil"], organe=organe, age="tous", stade=None)
        for vp in attendus:
            lignes.append({**base, "population": vp[1], "variable": vp[0],
                           "valeur": row[vp]})
    return pd.concat([autres, pd.DataFrame(lignes)], ignore_index=True)


# ── API publique ────────────────────────────────────────────────────────────────
def charger_long(dossier="data", fictif=False):
    """Ingestion complète (OECI + régional) → DataFrame long du contrat. Émet les
    comptes niveau HÔPITAL (nouveaux, prévus au contrat) en plus de l'aphp/ghu."""
    desc = _charger_descriptif()
    src = desc["sources"]
    lignes = _charger_oeci_long(dossier, src["oeci"], fictif)
    lignes += _charger_regional_long(dossier, src["regional"], fictif)

    long = pd.DataFrame(lignes, columns=LONG_COLS)
    # Pas de bourrage NaN : on n'émet pas les valeurs absentes/masquées.
    long = long[long["valeur"].notna()].copy()
    long = _reconstruire_survie_appareil(long)
    # Complétion à 0 (transform procédural de l'ancien pipeline) sur les périmètres
    # CONSOMMÉS par report_builder (comptes/délais aphp+ghu ; comptes régionaux).
    long = _completer_zeros(long, "DIM APHP", ["aphp", "ghu"], _ATTENDUS_DIM)
    long = _completer_zeros(long, "BN", ["aphp", "type_etab"], _ATTENDUS_BN)
    long = _filtrer_periode(long)
    # Dédup défensif (clé d'unicité du contrat) ; on garde la 1re occurrence.
    long = long.drop_duplicates(subset=[c for c in LONG_COLS if c != "valeur"])
    _alerter_derive(long, desc)
    return long[LONG_COLS].reset_index(drop=True)


def _alerter_derive(long, desc):
    """Avertit si des modalités sortent des « modalites_attendues » du YAML."""
    att = desc.get("modalites_attendues", {})
    ghu = set(att.get("ghu_codes", []))
    hop = set(att.get("hopitaux", []))
    typ = set(att.get("type_etab", []))
    niv = set(att.get("niveau_interne", []))
    inc_niv = set(long["niveau"]) - niv
    if inc_niv:
        print(f"  ⚠ Dérive niveau : {sorted(inc_niv)}")
    ents = {n: set(long[long["niveau"] == n]["entite"]) for n in long["niveau"].unique()}
    for n, attendu in (("ghu", ghu), ("hopital", hop), ("type_etab", typ)):
        inc = ents.get(n, set()) - attendu
        if inc:
            print(f"  ⚠ Dérive entité niveau {n} (hors modalités attendues) : {sorted(inc)[:10]}")


# ── Mapping hôpital → GHU (relocalisé ; utilisé par les pages inter-hôpitaux) ────
def _fichiers_oeci_pub(dossier, fictif):
    desc = _charger_descriptif()
    return _fichiers_oeci(dossier, desc["sources"]["oeci"]["fichier"], fictif)


def mapping_hopital_ghu(dossier="data", fictif=False):
    """{hôpital → code GHU} dérivé de « Survie globale » (GHU col 4, hôpital col 5),
    union des millésimes. Cf. pages de comparaison inter-hôpitaux (survie)."""
    mapping = {}
    for _, path in _fichiers_oeci_pub(dossier, fictif):
        wb = openpyxl.load_workbook(path, data_only=True)
        if "Survie globale" not in wb.sheetnames:
            wb.close(); continue
        ws = wb["Survie globale"]
        for r in range(5, ws.max_row + 1):
            agg, _ = _resoudre_niveau(ws.cell(r, 1).value, {"mode": "mots_cles"})
            if agg != "hopital":
                continue
            hop, code = ws.cell(r, 5).value, _code_ghu(ws.cell(r, 4).value)
            if hop not in (None, "") and code:
                mapping.setdefault(str(hop).strip(), code)
        wb.close()
    return mapping


def mapping_hopital_ghu_delais(dossier="data", fictif=False):
    """{hôpital → code GHU} NATIF de « Délais PEC » (GHU col 2 forme courte, hôpital
    col 3), union des millésimes. Cf. comparaison inter-hôpitaux (délais)."""
    desc = _charger_descriptif()
    conf_niv = desc["sources"]["oeci"]["feuilles"]["delais_pec"]["niveau"]
    mapping = {}
    for _, path in _fichiers_oeci_pub(dossier, fictif):
        wb = openpyxl.load_workbook(path, data_only=True)
        if "Délais PEC" not in wb.sheetnames:
            wb.close(); continue
        ws = wb["Délais PEC"]
        for r in range(3, ws.max_row + 1):
            agg, _ = _resoudre_niveau(ws.cell(r, 1).value, conf_niv)
            if agg != "hopital":
                continue
            hop, code = ws.cell(r, 3).value, _code_ghu(ws.cell(r, 2).value)
            if hop not in (None, "") and code:
                mapping.setdefault(str(hop).strip(), code)
        wb.close()
    return mapping
