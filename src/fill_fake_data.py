#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Remplit les gabarits de format Excel (OECI + régional) avec des données
fictives réalistes, au format STRICTEMENT identique.

Principe : les gabarits fournis dans ``data/`` contiennent déjà le squelette de
dimensions (colonnes Niveau, GHU, Hôpital, Appareil, Organe, Année, Classe age…).
Ce script NE TOUCHE PAS aux colonnes de dimension ni aux en-têtes (y compris les
en-têtes multi-lignes et cellules fusionnées de « Survie globale » et « Délais
PEC ») : il ne remplit que les colonnes de mesures, par-dessus le squelette
existant, puis écrit des copies suffixées ``_fictif`` (les gabarits restent
intacts).

Deux exceptions documentées :
  * « Origine géo » est livré DÉJÀ rempli dans le gabarit : ses 3 colonnes de
    valeurs sont régénérées en fictif (uniquement dans la copie).
  * « Survie globale » est livré ENTIÈREMENT vide (pas même le squelette de
    dimensions) : son squelette est reconstruit à partir des tuples de
    dimensions de « Délais PEC » (mêmes 5 dimensions), puis rempli.

Reproductibilité : seed fixe (np.random.seed(42)), série entièrement déterministe.

Série annuelle OECI : le fichier OECI ne porte pas de colonne année (l'année est
dans le NOM de fichier). On produit donc un fichier fictif PAR année de la
couverture du tableau de bord (``ANNEES``, 2019→2023). Pour ressembler à une vraie
chronique (et non à des tirages indépendants), une BASE est tirée une seule fois
par (feuille, appareil, organe, hôpital) puis chaque année en est dérivée par un
multiplicateur conjoncturel (``OECI_MULT`` : 2019 = référence, 2020 ≈ -10 % COVID,
2021 reprise partielle, 2022-2023 légère croissance) augmenté d'un petit bruit
±2,5 % propre à chaque (feuille, année). La survie s'améliore légèrement chaque
année (``OECI_SURV_DELTA`` ≈ +0,4 pt/an). Le même facteur appliqué à toutes les
colonnes de comptage d'une feuille préserve mécaniquement les invariants ``≤``.

Cohérence garantie :
  * OECI : somme Hôpital → GHU → AP-HP, et Organe → Appareil (les lignes
    agrégées sont calculées par sommation des feuilles `Hopital Organe`).
  * Régional : agrégation interne Organe→Appareil→Total par établissement, et
    onglet Total = « Age < 18 ans » + « Age >= 18 ans » cellule à cellule
    (jointure sur (Niveau, Finess, Appareil, Organe, Année)).
  * Sous-effectifs ≤ effectif total ; pourcentages ∈ [0, 100] ; survie I-III >
    IV et 1 an > 5 ans ; délais : médiane > 0 et moyenne ≳ médiane ; origine géo
    IDF/hors-IDF/International ≈ 100 % par groupe ; creux COVID 2020 (OECI ~ -10 %,
    régional ~ -8 %).
"""
import os
import numpy as np
import openpyxl

np.random.seed(42)

DATA = os.path.join(os.path.dirname(__file__), "..", "data")
OECI_IN = os.path.join(DATA, "indicateurs_oeci_2025_M12.xlsx")  # gabarit (structure de réf.)
PAT_IN = os.path.join(DATA, "canceroBR_16-25_Pat_13032026.xlsx")
PAT_OUT = os.path.join(DATA, "canceroBR_16-25_Pat_13032026_fictif.xlsx")
SEJ_IN = os.path.join(DATA, "canceroBR_16-25_Sej_13032026.xlsx")
SEJ_OUT = os.path.join(DATA, "canceroBR_16-25_Sej_13032026_fictif.xlsx")

# Mapping Hôpital -> GHU (nom complet OECI), relevé dans l'onglet « Effectifs recherche ».
HOSP2GHU = {
    # GHU Nord
    "Beaujon": "APHP.Nord-Université de Paris", "Bichat": "APHP.Nord-Université de Paris",
    "Bretonneau": "APHP.Nord-Université de Paris", "Lariboisière": "APHP.Nord-Université de Paris",
    "Louis Mourier": "APHP.Nord-Université de Paris", "Robert Debré": "APHP.Nord-Université de Paris",
    "Saint-Louis": "APHP.Nord-Université de Paris", "Villemin-Paul Doumer": "APHP.Nord-Université de Paris",
    # GHU PSL (Université Paris Saclay)
    "Ambroise Paré": "APHP.Université Paris Saclay", "Antoine Beclère": "APHP.Université Paris Saclay",
    "Bicêtre": "APHP.Université Paris Saclay", "Paul Brousse": "APHP.Université Paris Saclay",
    "Raymond Poincaré": "APHP.Université Paris Saclay", "Sainte-Perine": "APHP.Université Paris Saclay",
    # GHU Centre
    "Broca": "APHP.Centre-Université de Paris", "Cochin": "APHP.Centre-Université de Paris",
    "Corentin Celton": "APHP.Centre-Université de Paris", "Hegp": "APHP.Centre-Université de Paris",
    "Hôtel-Dieu": "APHP.Centre-Université de Paris", "Necker": "APHP.Centre-Université de Paris",
    # GHU Mondor
    "Emile Roux": "APHP.Hôpitaux Universitaires Henri-Mondor", "Georges Clémenceau": "APHP.Hôpitaux Universitaires Henri-Mondor",
    "Henri Mondor": "APHP.Hôpitaux Universitaires Henri-Mondor", "Joffre": "APHP.Hôpitaux Universitaires Henri-Mondor",
    # GHU PSSD
    "Avicenne": "APHP.Hôpitaux Universitaires Paris-Seine-Saint-Denis",
    "Jean Verdier": "APHP.Hôpitaux Universitaires Paris-Seine-Saint-Denis",
    "René Muret": "APHP.Hôpitaux Universitaires Paris-Seine-Saint-Denis",
    # GHU SUN
    "Armand Trousseau": "APHP.Sorbonne Université", "Charles Foix": "APHP.Sorbonne Université",
    "Pitie-Salpêtrière": "APHP.Sorbonne Université", "Rothschild": "APHP.Sorbonne Université",
    "Saint-Antoine": "APHP.Sorbonne Université", "Tenon": "APHP.Sorbonne Université",
}

# Facteurs de taille déterministes (gros sites pluridisciplinaires vs petits).
_HOSP_FACTOR = {h: round(float(f), 3) for h, f in
                zip(sorted(HOSP2GHU), 0.4 + 1.4 * np.random.rand(len(HOSP2GHU)))}


def hosp_factor(hop):
    return _HOSP_FACTOR.get(hop, 1.0)


_ORG_FACTOR = {}


def org_factor(app, org):
    """Facteur de fréquence par (appareil, organe), tiré une fois, stable."""
    key = (app, org)
    if key not in _ORG_FACTOR:
        _ORG_FACTOR[key] = float(np.random.uniform(0.3, 1.7))
    return _ORG_FACTOR[key]


def rint(x):
    return int(max(0, round(x)))


# ═══════════════════ série annuelle OECI : conjoncture & dérivation ═════════
# Couverture du tableau de bord. CONSTANTE de module : étendre la plage suffit
# (OECI_MULT / OECI_SURV_DELTA s'y adaptent automatiquement).
ANNEES = (2019, 2020, 2021, 2022, 2023)


def _build_oeci_mult(annees):
    """Multiplicateur de volume par année (chronique, pas tirages indépendants) :
    2019 = référence (1,0) ; 2020 ≈ -10 % (creux COVID) ; 2021 reprise partielle
    (~ +5 % vs 2020) ; ≥ 2022 légère croissance ~ +2,5 %/an. Avant 2019 : léger
    recul rétrospectif (~ -2 %/an)."""
    m = {}
    for a in annees:
        if a <= 2019:
            m[a] = round(0.98 ** (2019 - a), 4)
        elif a == 2020:
            m[a] = 0.90
        elif a == 2021:
            m[a] = 0.945
        else:
            m[a] = round(0.945 * 1.025 ** (a - 2021), 4)
    return m


def _build_oeci_surv_delta(annees):
    """Amélioration de survie cumulée vs 2019 : ~ +0,4 pt/an (bornée à 100 au
    moment de l'écriture). Garantit survie_2023 ≳ survie_2019."""
    return {a: round(0.4 * (a - 2019), 2) for a in annees}


OECI_MULT = _build_oeci_mult(ANNEES)
OECI_SURV_DELTA = _build_oeci_surv_delta(ANNEES)

# Caches de BASE (tirée une seule fois, seed 42) — réutilisés pour chaque année.
BASE_PATIENT = {}   # (app, org, hop) -> dict colonne->valeur
BASE_SEJOUR = {}
BASE_CHIR = {}
BASE_DELAIS = {}
BASE_SURVIE = {}
BASE_EFF = {}       # (ghu, hop, app, org, age) -> dict
BASE_ORIGINE = {}   # (niveau, ghu, hop) -> (total_base, poids origines)


def _yf(mult):
    """Facteur (feuille, année) = multiplicateur conjoncturel × bruit ±2,5 %.
    Appliqué à l'identique à TOUTES les colonnes de comptage d'une feuille, il
    préserve les relations ``sous-effectif ≤ total`` héritées de la base."""
    return mult * (1.0 + float(np.random.uniform(-0.025, 0.025)))


def scale_counts(base, mult):
    """Onglets de comptages purs (patient, séjour) : tout est volume."""
    f = _yf(mult)
    return {c: rint(v * f) for c, v in base.items()}


def scale_chir(base, mult):
    """Chirurgie : col 6 = volume (mis à l'échelle) ; 7..12 = taux % stables
    (petit bruit annuel, bornés [0, 100])."""
    f = _yf(mult)
    out = {6: rint(base[6] * f)}
    for c in range(7, 13):
        out[c] = round(min(100.0, max(0.0, base[c] * (1 + float(np.random.uniform(-0.02, 0.02))))), 1)
    return out


def scale_delais(base, mult):
    """Délais : Nb & Nb urg mis à l'échelle (urg ≤ nb préservé), % urg recalculé ;
    moyenne/médiane stables avec un MÊME bruit par bloc (préserve moyenne ≳ médiane
    et médiane > 0)."""
    f = _yf(mult)
    out = {}
    for c0 in (6, 11, 16, 21):
        nb = rint(base[c0] * f)
        urg = min(nb, rint(base[c0 + 1] * f))
        out[c0] = nb
        out[c0 + 1] = urg
        out[c0 + 2] = round(urg / nb * 100, 1) if nb else 0.0
        bn = 1 + float(np.random.uniform(-0.02, 0.02))
        out[c0 + 3] = round(max(0.2, base[c0 + 3] * bn), 1)
        out[c0 + 4] = round(max(0.1, base[c0 + 4] * bn), 1)
    return out


def scale_survie(base, mult, surv_delta):
    """Survie : Nb mis à l'échelle (Nvx ≤ Tous, IV ≤ I-III préservés) ; % décalés
    uniformément de +surv_delta (+ bruit ±0,15 pt) — le décalage uniforme préserve
    I-III > IV et 1 an > 5 ans, et fait croître la survie dans le temps."""
    f = _yf(mult)
    out = {}
    for c in (6, 8, 10, 12, 14, 16, 18, 20):
        out[c] = rint(base[c] * f)
    for c in (7, 9, 11, 13, 15, 17, 19, 21):
        out[c] = round(min(100.0, max(0.0, base[c] + surv_delta + float(np.random.uniform(-0.15, 0.15)))), 1)
    return out


def scale_eff(base, mult):
    """Effectifs recherche : patients & séjours mis à l'échelle (séjours ≥ patients
    préservé)."""
    f = _yf(mult)
    return {7: rint(base[7] * f), 8: rint(base[8] * f)}


def make_provider(cache, gen_fn, scale_fn, *scale_args):
    """Fournisseur de feuille pour une année : tire la BASE une seule fois (mise en
    cache au 1er passage) puis renvoie sa version dérivée pour l'année courante."""
    def provider(app, org, hop):
        key = (app, org, hop)
        if key not in cache:
            cache[key] = gen_fn(app, org, hop)
        return scale_fn(cache[key], *scale_args)
    return provider


def make_eff_provider(mult):
    def provider(ghu, hop, app, org, age):
        key = (ghu, hop, app, org, age)
        if key not in BASE_EFF:
            BASE_EFF[key] = gen_effectif_leaf(app, org, hop, age)
        return scale_eff(BASE_EFF[key], mult)
    return provider


# ───────────────────────── sélection des feuilles pour agrégation OECI ──────
def select_leaves(niv, ghu_col, hop, app, org, leaves):
    """leaves : liste de (ghu_full, hop, app, org, vals).
    Renvoie le sous-ensemble de feuilles agrégées par la ligne (niv, …)."""
    if niv == "GH Organe":
        return [v for (g, h, a, o, v) in leaves if g == ghu_col and a == app and o == org]
    if niv == "APHP Organe":
        return [v for (g, h, a, o, v) in leaves if a == app and o == org]
    if niv == "Hop Total":
        return [v for (g, h, a, o, v) in leaves if h == hop]
    if niv == "GH Total":
        return [v for (g, h, a, o, v) in leaves if g == ghu_col]
    if niv == "APHP Total":
        return [v for (g, h, a, o, v) in leaves]
    return []


def collect_rows(ws, hdr):
    """Retourne [(rowidx, niveau, ghu, hopital, appareil, organe), …] sur les
    lignes de données non vides (colonnes 1..5)."""
    out = []
    for r in range(hdr + 1, ws.max_row + 1):
        niv = ws.cell(r, 1).value
        if niv in (None, ""):
            continue
        out.append((r, niv, ws.cell(r, 2).value, ws.cell(r, 3).value,
                    ws.cell(r, 4).value, ws.cell(r, 5).value))
    return out


# ════════════════════════════ OECI ═════════════════════════════════════════
def origine_groups(ws):
    """Regroupe les lignes consécutives partageant (Niveau, GHU, Hôpital).
    Renvoie [(key, [rowidx…], [origine…])]. Structure identique d'une année à
    l'autre (même gabarit rechargé) → clés et indices de lignes stables."""
    groups = []
    for r in range(2, ws.max_row + 1):
        niv = ws.cell(r, 1).value
        if niv in (None, ""):
            continue
        key = (niv, ws.cell(r, 2).value, ws.cell(r, 3).value)
        orig = ws.cell(r, 4).value
        if groups and groups[-1][0] == key:
            groups[-1][1].append(r); groups[-1][2].append(orig)
        else:
            groups.append((key, [r], [orig]))
    return groups


def fill_origine_geo_year(ws, mult):
    """Origine géo pour une année : 3 colonnes Nb patients (5), Nb tot (6), % (7).
    Le total de référence et les proportions par origine sont tirés une seule fois
    par groupe (cache ``BASE_ORIGINE``), puis le total est mis à l'échelle par le
    multiplicateur annuel (+ bruit ±2,5 %), proportions conservées."""
    for key, rows, origs in origine_groups(ws):
        niv, ghu, hop = key
        if key not in BASE_ORIGINE:
            scale = hosp_factor(hop) if hop else (6.0 if "GH" in (niv or "") else 33.0)
            total_base = float(np.random.uniform(800, 3500) * max(scale, 0.5))
            ref = {"IDF": np.random.uniform(0.70, 0.82),
                   "Fr_notIDF": np.random.uniform(0.12, 0.25),
                   "International": np.random.uniform(0.003, 0.03)}
            w = np.array([ref.get(o, 0.1) for o in origs], float)
            w = w / w.sum()
            BASE_ORIGINE[key] = (total_base, w)
        total_base, w = BASE_ORIGINE[key]
        total = rint(total_base * mult * (1 + float(np.random.uniform(-0.025, 0.025))))
        parts = [rint(total * x) for x in w]
        diff = total - sum(parts)
        parts[int(np.argmax(parts))] += diff  # ajuste l'arrondi sur la plus grande part
        for r, part in zip(rows, parts):
            ws.cell(r, 5).value = part
            ws.cell(r, 6).value = total
            ws.cell(r, 7).value = round(part / total * 100, 1) if total else 0.0


def gen_patient_leaf(app, org, hop):
    base = rint(np.random.gamma(2.2, 18) * org_factor(app, org) * hosp_factor(hop) + 3)
    nvx = rint(base * np.random.uniform(0.40, 0.60))
    chir = rint(base * np.random.uniform(0.25, 0.55))
    nvx_chir = rint(min(chir, nvx) * np.random.uniform(0.5, 0.95))
    chimio = rint(base * np.random.uniform(0.20, 0.55))
    radio = rint(base * np.random.uniform(0.10, 0.40))
    sp = rint(base * np.random.uniform(0.05, 0.20))
    morts = rint(base * np.random.uniform(0.01, 0.07))
    morts_sp = rint(min(morts, sp) * np.random.uniform(0.3, 0.9))
    cart = rint(base * np.random.uniform(0, 0.03))
    greffe = rint(base * np.random.uniform(0, 0.04))
    return {6: base, 7: nvx, 8: chir, 9: nvx_chir, 10: chimio, 11: radio, 12: sp,
            13: morts, 14: morts_sp, 15: cart, 16: greffe}


def gen_sejour_leaf(app, org, hop):
    total = rint(np.random.gamma(2.4, 26) * org_factor(app, org) * hosp_factor(hop) + 4)
    chimio = rint(total * np.random.uniform(0.10, 0.45))
    chimio0 = rint(chimio * np.random.uniform(0.4, 0.9))
    chimio0_ped = rint(chimio0 * np.random.uniform(0, 0.15))
    radio = rint(total * np.random.uniform(0.05, 0.30))
    autres = rint(total * np.random.uniform(0.05, 0.25))
    chir = rint(total * np.random.uniform(0.10, 0.40))
    ped = rint(total * np.random.uniform(0, 0.12))
    sp = rint(total * np.random.uniform(0.02, 0.12))
    greffe = rint(total * np.random.uniform(0, 0.05))
    hemato = rint(total * np.random.uniform(0, 0.20))
    return {6: total, 7: chimio, 8: chimio0, 9: chimio0_ped, 10: radio, 11: autres,
            12: chir, 13: ped, 14: sp, 15: greffe, 16: hemato}


def gen_chirurgie_leaf(app, org, hop):
    nb = rint(np.random.gamma(2.0, 12) * org_factor(app, org) * hosp_factor(hop) + 2)
    rehosp30 = np.random.uniform(3, 12)
    rehosp90 = rehosp30 + np.random.uniform(1, 8)
    rechir30 = np.random.uniform(1, 6)
    rechir90 = rechir30 + np.random.uniform(0.5, 4)
    deces30 = np.random.uniform(0.3, 3.5)
    deces90 = deces30 + np.random.uniform(0.3, 3)
    return {6: nb,
            7: round(min(rehosp30, 100), 1), 8: round(min(rehosp90, 100), 1),
            9: round(min(rechir30, 100), 1), 10: round(min(rechir90, 100), 1),
            11: round(min(deces30, 100), 1), 12: round(min(deces90, 100), 1)}


def gen_delais_leaf(app, org, hop):
    """4 blocs de 5 colonnes : TOTAL(6-10), CHIRURGIE(11-15), MEDECINE(16-20),
    RADIOTHERAPIE(21-25) ; chaque bloc = Nb, Nb urg, %urg, Moyenne, Médiane."""
    base = rint(np.random.gamma(2.0, 14) * org_factor(app, org) * hosp_factor(hop) + 3)
    out = {}
    cols = {"TOTAL": 6, "CHIR": 11, "MED": 16, "RADIO": 21}
    frac = {"TOTAL": 1.0, "CHIR": np.random.uniform(0.3, 0.7),
            "MED": np.random.uniform(0.3, 0.7), "RADIO": np.random.uniform(0.2, 0.5)}
    med_ref = {"TOTAL": np.random.uniform(18, 35), "CHIR": np.random.uniform(20, 45),
               "MED": np.random.uniform(15, 30), "RADIO": np.random.uniform(25, 55)}
    for blk, c0 in cols.items():
        nb = rint(base * frac[blk])
        urg = rint(nb * np.random.uniform(0.05, 0.25))
        pct = round(urg / nb * 100, 1) if nb else 0.0
        med = max(1.0, med_ref[blk])
        moy = round(med * np.random.uniform(1.02, 1.20), 1)  # moyenne légèrement > médiane
        out[c0] = nb; out[c0 + 1] = urg; out[c0 + 2] = pct
        out[c0 + 3] = moy; out[c0 + 4] = round(med, 1)
    return out


def gen_survie_leaf(app, org, hop):
    """16 colonnes (6..21). Schéma : {Tous|Nvx} × {5ans|1an} × {I-III|IV}
    × {Nb, %}. I-III > IV ; 1 an > 5 ans ; Nvx ≤ Tous."""
    f = org_factor(app, org) * hosp_factor(hop)
    nb13_tous = rint(np.random.gamma(2.0, 14) * f + 3)
    nb4_tous = rint(nb13_tous * np.random.uniform(0.10, 0.35))
    nb13_nvx = rint(nb13_tous * np.random.uniform(0.40, 0.60))
    nb4_nvx = rint(nb4_tous * np.random.uniform(0.40, 0.60))
    s5_13 = np.random.uniform(55, 88)          # survie 5 ans, stade I-III
    s5_4 = np.random.uniform(8, min(40, s5_13 - 5))   # stade IV < I-III
    s1_13 = min(99.0, s5_13 + np.random.uniform(6, 14))   # 1 an > 5 ans
    s1_4 = min(99.0, s5_4 + np.random.uniform(8, 20))
    s1_4 = min(s1_4, s1_13 - 3)                # I-III > IV à 1 an aussi
    # colonnes : 6=Tous5a I-III nb,7=%, 8=IV nb,9=%, 10=Tous1a I-III nb,11=%,12=IV nb,13=%
    #            14..21 = idem « Nouveaux »
    def blk(nb13, nb4):
        return [nb13, round(s5_13, 1), nb4, round(s5_4, 1),
                nb13, round(s1_13, 1), nb4, round(s1_4, 1)]
    vals = blk(nb13_tous, nb4_tous) + blk(nb13_nvx, nb4_nvx)
    return {6 + i: v for i, v in enumerate(vals)}


def gen_effectif_leaf(app, org, hop, age):
    """Feuille « Effectifs recherche » (leaf = Hôpital-Organe-Appareil) :
    Nb patients (7) pondéré par classe d'âge, Nb séjours (8) ≥ patients."""
    pat = rint(np.random.gamma(1.6, 6) * org_factor(app, org) * hosp_factor(hop)
               * (1.0 if age == "Adultes" else 0.5 if age == "Gériatrie" else 0.2) + 1)
    sej = rint(pat * np.random.uniform(1.2, 3.5))
    return {7: pat, 8: sej}


def fill_count_sheet(ws, hdr, leaf_value, measure_cols):
    """Onglets de comptages purs (patient, séjour) : agrégats = sommes.
    ``leaf_value(app, org, hop)`` fournit la feuille déjà dérivée pour l'année."""
    rows = collect_rows(ws, hdr)
    vals = {}
    leaves = []
    for (r, niv, ghu, hop, app, org) in rows:
        if niv == "Hopital Organe":
            v = leaf_value(app, org, hop)
            vals[r] = v
            leaves.append((HOSP2GHU.get(hop), hop, app, org, v))
    for (r, niv, ghu, hop, app, org) in rows:
        if niv == "Hopital Organe":
            continue
        sel = select_leaves(niv, ghu, hop, app, org, leaves)
        vals[r] = {c: sum(v[c] for v in sel) for c in measure_cols}
    for r, v in vals.items():
        for c, x in v.items():
            ws.cell(r, c).value = x


def fill_rate_sheet(ws, hdr, leaf_value, count_col, rate_cols):
    """Onglet chirurgie : count_col sommé, rate_cols en moyenne pondérée par count_col.
    ``leaf_value(app, org, hop)`` fournit la feuille déjà dérivée pour l'année."""
    rows = collect_rows(ws, hdr)
    vals = {}
    leaves = []
    for (r, niv, ghu, hop, app, org) in rows:
        if niv == "Hopital Organe":
            v = leaf_value(app, org, hop)
            vals[r] = v
            leaves.append((HOSP2GHU.get(hop), hop, app, org, v))
    for (r, niv, ghu, hop, app, org) in rows:
        if niv == "Hopital Organe":
            continue
        sel = select_leaves(niv, ghu, hop, app, org, leaves)
        W = sum(v[count_col] for v in sel)
        agg = {count_col: W}
        for c in rate_cols:
            agg[c] = round(sum(v[c] * v[count_col] for v in sel) / W, 1) if W else 0.0
        vals[r] = agg
    for r, v in vals.items():
        for c, x in v.items():
            ws.cell(r, c).value = x


def fill_delais_sheet(ws, hdr, leaf_value):
    """4 blocs (TOTAL/CHIR/MED/RADIO). Nb & urg sommés ; %urg dérivé ;
    moyenne/médiane en moyenne pondérée par Nb.
    ``leaf_value(app, org, hop)`` fournit la feuille déjà dérivée pour l'année."""
    rows = collect_rows(ws, hdr)
    vals = {}
    leaves = []
    blocks = [6, 11, 16, 21]
    for (r, niv, ghu, hop, app, org) in rows:
        if niv == "Hopital Organe":
            v = leaf_value(app, org, hop)
            vals[r] = v
            leaves.append((HOSP2GHU.get(hop), hop, app, org, v))
    for (r, niv, ghu, hop, app, org) in rows:
        if niv == "Hopital Organe":
            continue
        sel = select_leaves(niv, ghu, hop, app, org, leaves)
        agg = {}
        for c0 in blocks:
            nb = sum(v[c0] for v in sel)
            urg = sum(v[c0 + 1] for v in sel)
            agg[c0] = nb
            agg[c0 + 1] = urg
            agg[c0 + 2] = round(urg / nb * 100, 1) if nb else 0.0
            if nb:
                agg[c0 + 3] = round(sum(v[c0 + 3] * v[c0] for v in sel) / nb, 1)
                agg[c0 + 4] = round(sum(v[c0 + 4] * v[c0] for v in sel) / nb, 1)
            else:
                agg[c0 + 3] = 0.0
                agg[c0 + 4] = 0.0
        vals[r] = agg
    for r, v in vals.items():
        for c, x in v.items():
            ws.cell(r, c).value = x


def fill_survie_sheet(ws_sur, ws_del, leaf_value):
    """« Survie globale » est vide : on reconstruit son squelette à partir des
    tuples de dimensions de « Délais PEC », puis on remplit. Colonnes survie :
    Niveau(1), Appareil patient(2), Organe patient(3), GHU(4), Hôpital(5).
    ``leaf_value(app, org, hop)`` fournit la feuille déjà dérivée pour l'année."""
    del_rows = collect_rows(ws_del, 2)  # (r, niv, ghu, hop, app, org)
    nb_cols = [6, 8, 10, 12, 14, 16, 18, 20]   # colonnes « Nb patients »
    pct_cols = [7, 9, 11, 13, 15, 17, 19, 21]  # colonnes « % survie »
    out = []   # (niv, ghu, hop, app, org)
    for (r, niv, ghu, hop, app, org) in del_rows:
        out.append((niv, ghu, hop, app, org))
    # écriture du squelette (lignes 5..) + génération/agrégation
    leaves = []
    rowvals = {}
    base_r = 5
    for i, (niv, ghu, hop, app, org) in enumerate(out):
        r = base_r + i
        ws_sur.cell(r, 1).value = niv
        ws_sur.cell(r, 2).value = app
        ws_sur.cell(r, 3).value = org
        ws_sur.cell(r, 4).value = ghu
        ws_sur.cell(r, 5).value = hop
        if niv == "Hopital Organe":
            v = leaf_value(app, org, hop)
            rowvals[r] = v
            leaves.append((HOSP2GHU.get(hop), hop, app, org, v))
    for i, (niv, ghu, hop, app, org) in enumerate(out):
        r = base_r + i
        if niv == "Hopital Organe":
            continue
        sel = select_leaves(niv, ghu, hop, app, org, leaves)
        agg = {}
        for c in nb_cols:
            agg[c] = sum(v[c] for v in sel)
        for c in pct_cols:
            W = sum(v[c - 1] for v in sel)  # pondération par le Nb du même bloc
            agg[c] = round(sum(v[c] * v[c - 1] for v in sel) / W, 1) if W else 0.0
        rowvals[r] = agg
    for r, v in rowvals.items():
        for c, x in v.items():
            ws_sur.cell(r, c).value = x


def fill_effectifs(ws, hdr, leaf_value):
    """Niveau hiérarchique avec dimension Classe age (col 6). Mesures : Nb
    patients (7), Nb séjours (8) ≥ patients. Leaf = 'Hôpital - Organe - Appareil'.
    ``leaf_value(ghu, hop, app, org, age)`` fournit la feuille dérivée pour l'année."""
    rows = []
    for r in range(hdr + 1, ws.max_row + 1):
        niv = ws.cell(r, 1).value
        if niv in (None, ""):
            continue
        rows.append((r, niv, ws.cell(r, 2).value, ws.cell(r, 3).value,
                     ws.cell(r, 4).value, ws.cell(r, 5).value, ws.cell(r, 6).value))
    vals = {}
    leaves = []  # (ghu, hop, app, org, age, pat, sej)
    for (r, niv, ghu, hop, app, org, age) in rows:
        if niv == "Hôpital - Organe - Appareil":
            v = leaf_value(ghu, hop, app, org, age)
            pat, sej = v[7], v[8]
            vals[r] = {7: pat, 8: sej}
            leaves.append((ghu, hop, app, org, age, pat, sej))

    def agg(pred):
        sub = [(p, s) for (g, h, a, o, ag, p, s) in leaves if pred(g, h, a, o, ag)]
        return {7: sum(p for p, s in sub), 8: sum(s for p, s in sub)}

    for (r, niv, ghu, hop, app, org, age) in rows:
        if niv == "Hôpital - Organe - Appareil":
            continue
        if niv == "GHU - Organe - Appareil":
            vals[r] = agg(lambda g, h, a, o, ag: g == ghu and a == app and o == org and ag == age)
        elif niv == "AP-HP - Organe - Appareil":
            vals[r] = agg(lambda g, h, a, o, ag: a == app and o == org and ag == age)
        elif niv == "Hopital":
            vals[r] = agg(lambda g, h, a, o, ag: h == hop and ag == age)
        elif niv == "GHU":
            vals[r] = agg(lambda g, h, a, o, ag: g == ghu and ag == age)
        elif niv == "AP-HP":
            vals[r] = agg(lambda g, h, a, o, ag: ag == age)
        else:
            vals[r] = {7: 0, 8: 0}
    for r, v in vals.items():
        for c, x in v.items():
            ws.cell(r, c).value = x


def oeci_out_path(annee):
    """Chemin du fichier annuel : l'année est portée par le NOM (convention OECI)."""
    return os.path.join(DATA, f"indicateurs_oeci_{annee}_M12_fictif.xlsx")


def fill_oeci_year(wb, annee):
    """Remplit un classeur OECI (gabarit rechargé) pour une année donnée, en
    dérivant la BASE partagée par le multiplicateur conjoncturel de l'année."""
    mult = OECI_MULT[annee]
    sd = OECI_SURV_DELTA[annee]
    fill_origine_geo_year(wb["Origine géo"], mult)
    fill_count_sheet(wb["Indicateurs patient"], 1,
                     make_provider(BASE_PATIENT, gen_patient_leaf, scale_counts, mult), list(range(6, 17)))
    fill_count_sheet(wb["Indicateurs séjour"], 1,
                     make_provider(BASE_SEJOUR, gen_sejour_leaf, scale_counts, mult), list(range(6, 17)))
    fill_rate_sheet(wb["Indicateurs chirurgie"], 1,
                    make_provider(BASE_CHIR, gen_chirurgie_leaf, scale_chir, mult), 6, list(range(7, 13)))
    fill_delais_sheet(wb["Délais PEC"], 2,
                      make_provider(BASE_DELAIS, gen_delais_leaf, scale_delais, mult))
    fill_survie_sheet(wb["Survie globale"], wb["Délais PEC"],
                      make_provider(BASE_SURVIE, gen_survie_leaf, scale_survie, mult, sd))
    fill_effectifs(wb["Effectifs recherche"], 1, make_eff_provider(mult))


def _find_niveau_row(ws, niveau, hdr=1):
    """Indice de la 1ʳᵉ ligne de niveau donné (ex. 'APHP Total'), ou None."""
    for r in range(hdr + 1, ws.max_row + 1):
        if ws.cell(r, 1).value == niveau:
            return r
    return None


def read_recap(wb):
    """Indicateurs AP-HP (toutes localisations) pour le récapitulatif inter-années :
    nb patients (Indicateurs patient · APHP Total · col 6) et survie 5 ans / 1 an
    stade I-III « Tous patients » (Survie globale · APHP Total · col 7 / 11)."""
    wp = wb["Indicateurs patient"]
    rp = _find_niveau_row(wp, "APHP Total")
    nb_pat = wp.cell(rp, 6).value if rp else None
    ws = wb["Survie globale"]
    rs = _find_niveau_row(ws, "APHP Total")
    surv5 = ws.cell(rs, 7).value if rs else None
    surv1 = ws.cell(rs, 11).value if rs else None
    return {"nb_patients": nb_pat, "survie_5ans_I_III": surv5, "survie_1an_I_III": surv1}


def validate_oeci(wb, annee):
    """Vérifie pour une année : agrégats à 0 écart, bornes, ordres de survie/délais.
    Renvoie la liste des anomalies (vide si tout est conforme)."""
    from collections import defaultdict
    issues = []

    # 1) Agrégation Hôpital → GHU → AP-HP et Organe → (APHP Organe) sur l'onglet patient.
    ws = wb["Indicateurs patient"]
    rows = collect_rows(ws, 1)
    cols = list(range(6, 17))
    total = {c: 0 for c in cols}
    by_ghu = defaultdict(lambda: {c: 0 for c in cols})
    by_org = defaultdict(lambda: {c: 0 for c in cols})
    for (r, niv, ghu, hop, app, org) in rows:
        if niv != "Hopital Organe":
            continue
        g = HOSP2GHU.get(hop)
        for c in cols:
            v = ws.cell(r, c).value or 0
            total[c] += v
            by_ghu[g][c] += v
            by_org[(app, org)][c] += v
        # bornes leaf : Nvx ≤ Nb, Patients chir ≤ Nb
        nb, nvx, chir = ws.cell(r, 6).value or 0, ws.cell(r, 7).value or 0, ws.cell(r, 8).value or 0
        if nvx > nb or chir > nb:
            issues.append(f"patient borne leaf r{r}: nvx={nvx} chir={chir} > nb={nb}")
    for (r, niv, ghu, hop, app, org) in rows:
        if niv == "APHP Total":
            for c in cols:
                if (ws.cell(r, c).value or 0) != total[c]:
                    issues.append(f"patient APHP Total col{c}: {ws.cell(r, c).value} ≠ Σleaves {total[c]}")
        elif niv == "GH Total":
            for c in cols:
                if (ws.cell(r, c).value or 0) != by_ghu[ghu][c]:
                    issues.append(f"patient GH Total {ghu} col{c}: écart")
        elif niv == "APHP Organe":
            for c in cols:
                if (ws.cell(r, c).value or 0) != by_org[(app, org)][c]:
                    issues.append(f"patient APHP Organe {app}/{org} col{c}: écart")

    # 2) Survie AP-HP : I-III > IV (5 ans) et 1 an > 5 ans (I-III).
    ws = wb["Survie globale"]
    rs = _find_niveau_row(ws, "APHP Total")
    if rs:
        s5_13, s5_4 = ws.cell(rs, 7).value, ws.cell(rs, 9).value
        s1_13 = ws.cell(rs, 11).value
        if not (s5_13 > s5_4):
            issues.append(f"survie AP-HP : 5 ans I-III ({s5_13}) ≤ IV ({s5_4})")
        if not (s1_13 > s5_13):
            issues.append(f"survie AP-HP : 1 an ({s1_13}) ≤ 5 ans ({s5_13}) (I-III)")

    # 3) Délais AP-HP : médiane > 0 et moyenne ≳ médiane sur chaque bloc.
    ws = wb["Délais PEC"]
    rd = _find_niveau_row(ws, "APHP Total", hdr=2)
    if rd:
        for c0 in (6, 11, 16, 21):
            moy, med = ws.cell(rd, c0 + 3).value, ws.cell(rd, c0 + 4).value
            if not (med > 0):
                issues.append(f"délais AP-HP bloc col{c0}: médiane {med} ≤ 0")
            if moy + 0.05 < med:
                issues.append(f"délais AP-HP bloc col{c0}: moyenne {moy} < médiane {med}")
    return issues


def fill_oeci():
    """Produit la SÉRIE ANNUELLE OECI (un fichier fictif par année de ``ANNEES``).
    La BASE est tirée une seule fois (au 1er passage, cache ``BASE_*``) puis chaque
    année en est dérivée. Renvoie le récapitulatif inter-années [(annee, recap)…]."""
    print(f"OECI : série annuelle {ANNEES[0]}→{ANNEES[-1]} (base tirée une seule fois)…")
    recap = []
    all_issues = {}
    for annee in ANNEES:
        wb = openpyxl.load_workbook(OECI_IN)  # gabarit rechargé : structure intacte
        fill_oeci_year(wb, annee)
        issues = validate_oeci(wb, annee)
        all_issues[annee] = issues
        out = oeci_out_path(annee)
        wb.save(out)
        recap.append((annee, read_recap(wb)))
        flag = "OK" if not issues else f"{len(issues)} anomalie(s)"
        print(f"  → {annee} (×{OECI_MULT[annee]:.3f}) : {os.path.basename(out)}  [{flag}]")
    _print_recap(recap, all_issues)
    return recap


def _print_recap(recap, all_issues):
    """Affiche le récapitulatif inter-années AP-HP + le bilan de validation."""
    print("\n  Récapitulatif inter-années — AP-HP (toutes localisations) :")
    print(f"    {'Année':>6} | {'Nb patients':>12} | {'Survie 5 ans I-III':>18} | {'Survie 1 an I-III':>17}")
    print("    " + "-" * 62)
    prev_nb = prev_surv = None
    for annee, rc in recap:
        nb, s5 = rc["nb_patients"], rc["survie_5ans_I_III"]
        d_nb = "" if prev_nb is None else f"({(nb / prev_nb - 1) * 100:+.1f} %)"
        d_s = "" if prev_surv is None else f"({s5 - prev_surv:+.2f} pt)"
        print(f"    {annee:>6} | {nb:>12,} {d_nb:>9} | {s5:>10} {d_s:>7} | {rc['survie_1an_I_III']:>17}")
        prev_nb, prev_surv = nb, s5
    # contrôles transverses
    survs = [rc["survie_5ans_I_III"] for _, rc in recap]
    nbs = [rc["nb_patients"] for _, rc in recap]
    print("\n  Contrôles :")
    print(f"    creux COVID 2020 : nb min en {recap[nbs.index(min(nbs))][0]} "
          f"({'OK' if recap[nbs.index(min(nbs))][0] == 2020 else 'À VÉRIFIER'})")
    print(f"    survie croissante {recap[0][0]}→{recap[-1][0]} : "
          f"{survs[0]} → {survs[-1]} "
          f"({'OK' if survs[-1] >= survs[0] and survs == sorted(survs) else 'À VÉRIFIER'})")
    tot_issues = sum(len(v) for v in all_issues.values())
    print(f"    agrégats/bornes : {'OK — 0 anomalie' if tot_issues == 0 else f'{tot_issues} anomalie(s)'}")
    for annee, iss in all_issues.items():
        for msg in iss[:5]:
            print(f"      [{annee}] {msg}")


# ════════════════════════════ Régional ═════════════════════════════════════
# colonnes (1-based) : Pat -> dims 1..8, mesures 9..14 ; Sej -> dims 1..8, mesures 9..20
PAT_MEAS = list(range(9, 15))
SEJ_MEAS = list(range(9, 21))
# offset d'en-tête : Pat = 1 ligne, Sej = 2 lignes
# Conjoncture par année : base 2019 = 1,0, creux COVID 2020 (~ -8 %) puis légère reprise.
COVID = {2016: 0.95, 2017: 0.96, 2018: 0.98, 2019: 1.0, 2020: 0.92, 2021: 0.96,
         2022: 0.98, 2023: 1.0, 2024: 1.01, 2025: 1.02}


def _finess_factor(cache, finess, statut):
    if finess not in cache:
        base = {"CLCC": 2.2, "CHR/U": 1.6, "CH": 0.9, "PSPH/EBNL": 0.8, "Privé": 0.7}.get(statut, 1.0)
        cache[finess] = base * float(np.random.uniform(0.5, 1.5))
    return cache[finess]


def gen_pat_leaf(scale, age_sheet):
    """age_sheet ∈ {'<18','>=18'}. Renvoie dict mesure->val pour une ligne feuille."""
    if age_sheet == "<18":
        base = rint(np.random.gamma(1.4, 3) * scale + 0)
    else:
        base = rint(np.random.gamma(1.8, 9) * scale + 0)
    if base == 0:
        return {c: 0 for c in PAT_MEAS}
    nvx = rint(base * np.random.uniform(0.4, 0.65))
    sup75 = 0 if age_sheet == "<18" else rint(base * np.random.uniform(0.1, 0.4))
    chir = rint(base * np.random.uniform(0.2, 0.55))
    chimio = rint(base * np.random.uniform(0.15, 0.5))
    radio = rint(base * np.random.uniform(0.05, 0.35))
    return {9: base, 10: nvx, 11: sup75, 12: chir, 13: chimio, 14: radio}


def gen_sej_leaf(scale, age_sheet):
    if age_sheet == "<18":
        total = rint(np.random.gamma(1.5, 4) * scale)
    else:
        total = rint(np.random.gamma(2.0, 12) * scale)
    if total == 0:
        return {c: 0 for c in SEJ_MEAS}
    # composantes ≤ total
    chimio0 = rint(total * np.random.uniform(0.05, 0.3))
    radio0 = rint(total * np.random.uniform(0.02, 0.2))
    chir0 = rint(total * np.random.uniform(0.05, 0.25))
    autres0 = rint(total * np.random.uniform(0.05, 0.25))
    chir_p = rint(total * np.random.uniform(0.05, 0.3))
    j_chir = rint(chir_p * np.random.uniform(2, 8))       # journées ≥ séjours
    autres_p = rint(total * np.random.uniform(0.05, 0.25))
    j_autres = rint(autres_p * np.random.uniform(2, 8))
    pallia = rint(total * np.random.uniform(0.01, 0.1))
    journees = rint(total * np.random.uniform(2.5, 7))    # journées totales ≥ séjours
    non_chain = rint(total * np.random.uniform(0, 0.1))
    return {9: chimio0, 10: radio0, 11: chir0, 12: autres0,
            13: chir_p, 14: j_chir, 15: autres_p, 16: j_autres,
            17: pallia, 18: total, 19: journees, 20: non_chain}


def _row_key(ws, r):
    """Clé d'agrégation âge : (Niveau, Finess, Appareil, Organe, Année)."""
    return (ws.cell(r, 1).value, ws.cell(r, 3).value, ws.cell(r, 6).value,
            ws.cell(r, 7).value, ws.cell(r, 8).value)


def fill_regional(path_in, path_out, gen_leaf, meas_cols, hdr):
    print(f"Régional : lecture de {os.path.basename(path_in)} (peut être long)…")
    wb = openpyxl.load_workbook(path_in)
    sheet_vals = {}   # sheetname -> {rowidx: {col: val}}
    age_dicts = {}    # sheetname -> {key: {col: val}}  (toutes lignes)
    fcache = {}
    # 1) feuilles d'âge : génère feuilles 'Organe' puis agrège Appareil/Total
    for sname, tag in [("Age < 18 ans", "<18"), ("Age >= 18 ans", ">=18")]:
        ws = wb[sname]
        vals = {}
        leaf_idx = {}  # (finess, appareil, organe, année) -> dict
        for r in range(hdr + 1, ws.max_row + 1):
            niv = ws.cell(r, 1).value
            if niv in (None, ""):
                continue
            if niv == "Organe":
                finess = ws.cell(r, 3).value
                statut = ws.cell(r, 5).value
                annee = ws.cell(r, 8).value
                scale = _finess_factor(fcache, finess, statut) * COVID.get(annee, 1.0)
                v = gen_leaf(scale, tag)
                vals[r] = v
                leaf_idx[(finess, ws.cell(r, 6).value, ws.cell(r, 7).value, annee)] = v
        # agrégats Appareil (Σ organes même finess/appareil/année) et Total (Σ même finess/année)
        from collections import defaultdict
        by_app = defaultdict(lambda: {c: 0 for c in meas_cols})
        by_tot = defaultdict(lambda: {c: 0 for c in meas_cols})
        for (finess, app, org, annee), v in leaf_idx.items():
            ka = (finess, app, annee)
            kt = (finess, annee)
            for c in meas_cols:
                by_app[ka][c] += v[c]
                by_tot[kt][c] += v[c]
        for r in range(hdr + 1, ws.max_row + 1):
            niv = ws.cell(r, 1).value
            if niv == "Appareil":
                vals[r] = dict(by_app[(ws.cell(r, 3).value, ws.cell(r, 6).value, ws.cell(r, 8).value)])
            elif niv == "Total":
                vals[r] = dict(by_tot[(ws.cell(r, 3).value, ws.cell(r, 8).value)])
        sheet_vals[sname] = vals
        # index par clé pour la somme d'âges
        ad = {}
        for r, v in vals.items():
            ad[_row_key(ws, r)] = v
        age_dicts[sname] = ad
        print(f"  {sname}: {len(vals)} lignes remplies")
    # 2) onglet Total = Σ des deux feuilles d'âge, cellule à cellule
    ws = wb["Total"]
    d1 = age_dicts["Age < 18 ans"]
    d2 = age_dicts["Age >= 18 ans"]
    vals = {}
    miss = 0
    for r in range(hdr + 1, ws.max_row + 1):
        niv = ws.cell(r, 1).value
        if niv in (None, ""):
            continue
        k = _row_key(ws, r)
        v1 = d1.get(k); v2 = d2.get(k)
        if v1 is None and v2 is None:
            miss += 1
            vals[r] = {c: 0 for c in meas_cols}
            continue
        vals[r] = {c: (v1[c] if v1 else 0) + (v2[c] if v2 else 0) for c in meas_cols}
    sheet_vals["Total"] = vals
    print(f"  Total: {len(vals)} lignes (clé absente des 2 âges : {miss})")
    # 3) écriture
    for sname, vals in sheet_vals.items():
        ws = wb[sname]
        for r, v in vals.items():
            for c, x in v.items():
                ws.cell(r, c).value = x
    wb.save(path_out)
    print(f"  → écrit : {path_out}")


def main():
    fill_oeci()
    fill_regional(PAT_IN, PAT_OUT, gen_pat_leaf, PAT_MEAS, hdr=1)
    fill_regional(SEJ_IN, SEJ_OUT, gen_sej_leaf, SEJ_MEAS, hdr=2)
    print("\nTerminé.")


if __name__ == "__main__":
    main()
